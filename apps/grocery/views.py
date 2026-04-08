from collections import defaultdict
from decimal import Decimal

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.engine.calculation import format_quantity
from apps.engine.models import EventIngredient
from shared.exports.excel_service import create_workbook, workbook_to_bytes
from shared.permissions import IsTenantScopedJWT


CATEGORY_ORDER = [
    'GROCERY',
    'DISPOSABLE',
    'VEGETABLE',
    'FRUIT',
    'RENTAL',
    'CHICKEN',
    'BEEF',
    'MUTTON',
    'FISH',
    'MEAT',
    'OTHER',
]


def _category_sort_key(cat: str) -> int:
    upper = cat.upper()
    try:
        return CATEGORY_ORDER.index(upper)
    except ValueError:
        return 99


def _get_events_for_date(date_str: str, tenant_id: str):
    """Return confirmed events (for this tenant) whose event_date matches date_str."""
    from apps.events.models import Event
    return Event.objects.filter(
        tenant_id=tenant_id,
        event_date=date_str,
        status__in=['CONFIRMED', 'IN_PROGRESS'],
        is_deleted=False,
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsTenantScopedJWT])
def grocery_list(request):
    """
    GET /api/app/<slug>/grocery/?date=YYYY-MM-DD[&category=GROCERY&category=VEGETABLE…]

    Returns aggregated ingredients across all confirmed events on that date
    for the current tenant.
    """
    date_str = request.query_params.get('date', '')
    if not date_str:
        return Response({'error': 'date is required (YYYY-MM-DD)'}, status=400)

    category_filter = [c.upper() for c in request.query_params.getlist('category') if c]
    events, ingredients = _build_ingredients(date_str, category_filter, request.tenant_id)

    events_data = [
        {
            'id': str(ev.id),
            'event_id': ev.event_code,
            'event_name': ev.event_type or ev.event_code,
            'client_name': ev.customer_name,
            'guest_count': ev.guest_count,
        }
        for ev in events
    ]

    serialized = [
        {**item, 'quantity': str(item['quantity'])}
        for item in ingredients
    ]

    return Response({
        'events': events_data,
        'ingredients': serialized,
        'total_ingredients': len(serialized),
        'total_events': len(events_data),
    })


def _build_ingredients(date_str: str, category_filter: list[str], tenant_id: str) -> tuple[list, list]:
    """Shared aggregation used by both the list view and export."""
    events = _get_events_for_date(date_str, tenant_id)
    event_ids = list(events.values_list('id', flat=True))

    qs = EventIngredient.objects.filter(event_id__in=event_ids, tenant_id=tenant_id)
    if category_filter:
        qs = qs.filter(category__in=category_filter)

    totals: dict[tuple, dict] = defaultdict(lambda: {
        'total': Decimal('0'),
        'unit': '',
        'category': '',
    })
    for row in qs:
        key = (row.ingredient_name, row.unit)
        totals[key]['total'] += row.total_quantity
        totals[key]['unit'] = row.unit
        totals[key]['category'] = row.category

    ingredients = []
    for (name, _unit), data in totals.items():
        if data['total'] < Decimal('0.001'):
            continue  # skip negligible quantities
        qty, unit = format_quantity(data['total'], data['unit'])
        ingredients.append({
            'ingredient_name': name,
            'quantity': qty,
            'unit': unit,
            'category': data['category'],
        })
    ingredients.sort(key=lambda x: (_category_sort_key(x['category']), x['ingredient_name'].lower()))
    return list(events), ingredients


CATEGORY_LABELS = {
    'GROCERY': 'Grocery', 'DISPOSABLE': 'Disposable', 'VEGETABLE': 'Vegetable',
    'FRUIT': 'Fruit', 'RENTAL': 'Rental', 'CHICKEN': 'Chicken', 'BEEF': 'Beef',
    'MUTTON': 'Mutton', 'FISH': 'Fish', 'MEAT': 'Meat', 'OTHER': 'Other',
}


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsTenantScopedJWT])
def export_excel(request):
    """GET /api/app/<slug>/grocery/export/excel/?date=YYYY-MM-DD[&category=…]"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    date_str = request.query_params.get('date', '')
    if not date_str:
        return HttpResponse('date is required', status=400)

    category_filter = [c.upper() for c in request.query_params.getlist('category') if c]
    events, ingredients = _build_ingredients(date_str, category_filter, request.tenant_id)

    wb = create_workbook()
    ws = wb.active
    ws.title = 'Grocery List'

    # ── Title row ──────────────────────────────────────────────
    event_names = ', '.join(ev.customer_name for ev in events) or 'No events'
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Grocery List — {date_str}  |  {event_names}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # ── Header row ─────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='1C3355')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    headers = ['#', 'Ingredient', 'Quantity', 'Unit']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center' if col != 2 else 'left', vertical='center')

    # ── Data rows grouped by category ──────────────────────────
    thin = Border(
        bottom=Side(style='thin', color='E2E8F0'),
    )
    cat_fill = PatternFill('solid', fgColor='EFF6FF')
    cat_font = Font(bold=True, size=10, color='185FA5')

    row = 3
    serial = 1
    current_cat = None

    for item in ingredients:
        cat = item['category'].upper()
        if cat != current_cat:
            # Category header
            ws.merge_cells(f'A{row}:D{row}')
            label = CATEGORY_LABELS.get(cat, cat.title())
            ws[f'A{row}'] = label
            ws[f'A{row}'].fill = cat_fill
            ws[f'A{row}'].font = cat_font
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 18
            row += 1
            current_cat = cat
            serial = 1

        qty = float(item['quantity'])
        qty_str = f'{qty:g}'  # removes trailing zeros: 500.0→"500", 0.1→"0.1"

        ws.cell(row=row, column=1, value=serial).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=item['ingredient_name'])
        ws.cell(row=row, column=3, value=qty_str).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4, value=item['unit'].upper()).alignment = Alignment(horizontal='left')

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = thin

        row += 1
        serial += 1

    # ── Column widths ──────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10

    content = workbook_to_bytes(wb)
    response = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="grocery-{date_str}.xlsx"'
    return response
